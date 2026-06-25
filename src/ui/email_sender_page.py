"""
ZUGZWANG - Email Sender Page
Fully functional SMTP email sender with threading, attachments, persistence, and progress tracking.
Aesthetic 3.0: Senior Grade macOS Design.
"""

from __future__ import annotations

import html
import hashlib
import os
import smtplib
import ssl
import time
import threading
import socket
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

from PySide6.QtCore import Qt, Signal, QObject, QSize
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QFrame, QLabel,
    QLineEdit, QTextEdit, QFileDialog, QSizePolicy, QStackedWidget, QTextBrowser,
    QPlainTextEdit, QScrollArea, QPushButton, QCompleter, QDialog
)
from .components import StatCard, SectionCard, MacSwitch
from qfluentwidgets import (
    InfoBadge, ProgressBar,
    PushButton, PrimaryPushButton, TransparentPushButton, ToolButton,
    ElevatedCardWidget, FluentIcon, LineEdit, PlainTextEdit,
    TransparentToolButton,
    ScrollArea, RoundMenu, Action, IconWidget
)
from PySide6.QtGui import QTextCharFormat, QColor, QTextCursor

from ..core.events import event_bus
from .theme import Theme
from ..core.config import config_manager
from ..core.i18n import get_language, tr


class _SendSignals(QObject):
    log = Signal(str, str)        # message, level
    progress = Signal(int, int)   # current, total
    finished = Signal(int, int, bool) # sent, failed, was_aborted
    error = Signal(str)           # global error


from PySide6.QtWidgets import QListWidget, QListWidgetItem, QMenu, QAbstractItemView, QStyledItemDelegate, QStyleOptionViewItem, QApplication, QStyle
from PySide6.QtGui import QPainter, QCursor, QGuiApplication
import re

_EMAIL_PROFILE_RE = re.compile(r"^[^@\s]+@[^@\s]+\.com$", re.IGNORECASE)

class RecipientItem(QListWidgetItem):
    def __init__(self, email: str, parent=None):
        super().__init__(parent)
        self.email = email
        self.status = "pending" # pending, sending, success, failed
        self.setText(email)
        self.setSizeHint(QSize(0, 24))
        self.setFlags(self.flags() | Qt.ItemIsEditable)

    def setData(self, role, value):
        super().setData(role, value)
        if role in (Qt.DisplayRole, Qt.EditRole):
            self.email = str(value or "").strip()


class RecipientDelegate(QStyledItemDelegate):
    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index):
        rect = option.rect
        row_background = QColor("#1A1A1A")
        separator_color = QColor("#2C2C2E")

        if option.state & QStyle.State_Editing:
            painter.save()
            painter.setRenderHint(QPainter.Antialiasing)
            painter.fillRect(rect, row_background)
            painter.setBrush(QColor(10, 132, 255, 40) if option.state & QStyle.State_Selected else QColor(255, 255, 255, 10))
            painter.setPen(Qt.NoPen)
            painter.drawRoundedRect(rect, 6, 6)
            painter.setPen(separator_color)
            painter.drawLine(rect.bottomLeft(), rect.bottomRight())
            painter.restore()
            return

        painter.save()
        painter.setRenderHint(QPainter.Antialiasing)
        painter.fillRect(rect, row_background)

        # Hover / Selected bg
        if option.state & QStyle.State_Selected:
            painter.setBrush(QColor(10, 132, 255, 40))
        elif option.state & QStyle.State_MouseOver:
            painter.setBrush(QColor(255, 255, 255, 10))
        else:
            painter.setBrush(Qt.NoBrush)
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(rect, 6, 6)

        email = index.data(Qt.DisplayRole)
        if not email:
            painter.restore()
            return
            
        painter.setPen(QColor("#FFFFFF"))
        font = option.font
        font.setFamily("SF Mono")
        font.setPointSize(9.0)
        
        # Cross out already sent
        # Attempt to find history on the widget itself or its parent (EmailSenderPage)
        history = []
        if option.widget and hasattr(option.widget, "_current_sent_emails"):
            history = option.widget._current_sent_emails()
        elif option.widget and option.widget.parent() and hasattr(option.widget.parent(), "_current_sent_emails"):
            history = option.widget.parent()._current_sent_emails()
            
        if email.lower() in history:
            font.setStrikeOut(True)
            painter.setPen(QColor("#32D74B")) # Green Success Color
        
        painter.setFont(font)
        painter.setClipRect(rect.adjusted(36, 0, -12, 0))
        painter.drawText(
            rect.adjusted(36, 0, -12, 0),
            Qt.AlignLeft | Qt.AlignVCenter | Qt.TextSingleLine,
            email,
        )
        painter.setClipping(False)
        
        # Drag Handle (Left)
        painter.setPen(QColor("#636366"))
        font.setPointSize(10)
        painter.setFont(font)
        painter.drawText(rect.adjusted(12, 0, 0, 0), Qt.AlignLeft | Qt.AlignVCenter, "≡")
        painter.setPen(separator_color)
        painter.drawLine(rect.bottomLeft(), rect.bottomRight())

        painter.restore()

    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        editor.setFrame(False)
        editor.setClearButtonEnabled(True)
        editor.setAutoFillBackground(True)
        editor.setAttribute(Qt.WA_OpaquePaintEvent, True)
        editor.setStyleSheet("""
            QLineEdit {
                background: #1A1A1A;
                border: 1px solid rgba(10, 132, 255, 0.75);
                border-radius: 6px;
                color: #FFFFFF;
                padding: 0 8px;
                selection-background-color: rgba(10, 132, 255, 140);
                selection-color: #FFFFFF;
            }
        """)
        font = option.font
        font.setFamily("SF Mono")
        font.setPointSizeF(9.0)
        editor.setFont(font)
        return editor

    def setEditorData(self, editor, index):
        editor.setText(index.data(Qt.EditRole) or index.data(Qt.DisplayRole) or "")
        editor.selectAll()

    def setModelData(self, editor, model, index):
        model.setData(index, editor.text().strip(), Qt.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect.adjusted(34, 2, -10, -2))


class RecipientListWidget(QListWidget):
    items_changed = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            QListWidget {
                background: #1A1A1A;
                border: 1px solid #3A3A3C;
                border-radius: 10px;
                padding: 10px;
                outline: none;
            }
            QListWidget::item { border-bottom: 1px solid #2C2C2E; }
        """)
        self.setDragDropMode(QAbstractItemView.InternalMove)
        self.setDefaultDropAction(Qt.MoveAction)
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.setItemDelegate(RecipientDelegate(self))
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)
        
        # Track drag changes
        self.model().rowsMoved.connect(lambda: self.items_changed.emit())
        self.model().rowsInserted.connect(lambda: self.items_changed.emit())
        self.model().rowsRemoved.connect(lambda: self.items_changed.emit())
        self.itemChanged.connect(self._on_item_changed)

    def _on_item_changed(self, item):
        cleaned = item.text().strip()
        item.email = cleaned
        if item.text() != cleaned:
            self.blockSignals(True)
            item.setText(cleaned)
            self.blockSignals(False)
        self.items_changed.emit()

    def _show_context_menu(self, pos):
        item = self.itemAt(pos)
        selected = self.selectedItems()
        
        copy_this = None
        move_top_act = None
        copy_sel = None
        del_act = None

        from qfluentwidgets import RoundMenu, Action, FluentIcon
        menu = RoundMenu(parent=self)

        # Item-specific actions
        if item:
            copy_this = Action(FluentIcon.COPY, "Copy This Email", self)
            menu.addAction(copy_this)
            move_top_act = Action(FluentIcon.UP, "Move to Top", self)
            menu.addAction(move_top_act)
            menu.addSeparator()

        # Selection actions
        if len(selected) > 1:
            copy_sel = Action(FluentIcon.COPY, f"Copy Selected ({len(selected)})", self)
            menu.addAction(copy_sel)

        copy_all = Action(FluentIcon.COPY, "Copy All Queue", self)
        menu.addAction(copy_all)
        menu.addSeparator()

        # Multi-delete if selected
        if len(selected) > 0:
            del_act = Action(FluentIcon.DELETE, f"Remove {'Selection' if len(selected) > 1 else 'Email'}", self)
            menu.addAction(del_act)

        action = menu.exec(self.mapToGlobal(pos))
        
        from PySide6.QtGui import QGuiApplication
        clipboard = QGuiApplication.clipboard()

        if item and action == copy_this:
            clipboard.setText(item.email)
        elif copy_sel and action == copy_sel:
            text = "\n".join(it.email for it in selected)
            clipboard.setText(text)
        elif action == copy_all:
            text = "\n".join(self.item(i).email for i in range(self.count()))
            clipboard.setText(text)
        elif action == move_top_act:
            row = self.row(item)
            if row > 0:
                self.takeItem(row)
                self.insertItem(0, item)
        elif del_act and action == del_act:
            for it in selected:
                self.takeItem(self.row(it))



class ManualRecipientDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.recipient_email = ""

        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setFixedWidth(380)

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)

        container = QFrame(self)
        container.setObjectName("ManualRecipientDialog")
        container.setStyleSheet("""
            QFrame#ManualRecipientDialog {
                background: #1C1C1E;
                border: 1px solid #3A3A3C;
                border-radius: 16px;
            }
        """)
        root.addWidget(container)

        layout = QVBoxLayout(container)
        layout.setContentsMargins(24, 22, 24, 22)
        layout.setSpacing(14)

        title = QLabel("Add Recipient")
        title.setStyleSheet(
            "color: #FFFFFF; font-family: 'PT Root UI', sans-serif; "
            "font-size: 18px; font-weight: 700; background: transparent; border: none;"
        )
        layout.addWidget(title)

        subtitle = QLabel("Type one email address and add it directly to the queue.")
        subtitle.setWordWrap(True)
        subtitle.setStyleSheet(
            "color: #8E8E93; font-family: 'PT Root UI', sans-serif; "
            "font-size: 12px; background: transparent; border: none;"
        )
        layout.addWidget(subtitle)

        field_label = QLabel("RECIPIENT EMAIL")
        field_label.setStyleSheet(
            "color: #636366; font-family: 'PT Root UI', sans-serif; "
            "font-size: 10px; font-weight: 600; letter-spacing: 1px; "
            "background: transparent; border: none;"
        )
        layout.addWidget(field_label)

        self._email_input = LineEdit()
        self._email_input.setPlaceholderText("name@example.com")
        self._email_input.setFixedHeight(42)
        self._email_input.setStyleSheet("""
            QLineEdit {
                background: #2C2C2E;
                border: 1px solid #3A3A3C;
                border-radius: 10px;
                color: #FFFFFF;
                font-family: 'PT Root UI', sans-serif;
                font-size: 13px;
                padding: 0 12px;
            }
            QLineEdit:focus {
                border: 1px solid #0A84FF;
            }
        """)
        self._email_input.returnPressed.connect(self._submit)
        layout.addWidget(self._email_input)

        self._error_label = QLabel("")
        self._error_label.setWordWrap(True)
        self._error_label.setStyleSheet(
            "color: #FF9F0A; font-family: 'PT Root UI', sans-serif; "
            "font-size: 11px; background: transparent; border: none;"
        )
        self._error_label.hide()
        layout.addWidget(self._error_label)

        buttons = QHBoxLayout()
        buttons.setSpacing(10)

        cancel_btn = QPushButton("CANCEL")
        cancel_btn.setFixedHeight(38)
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background: #2C2C2E;
                border: 1px solid #3A3A3C;
                border-radius: 10px;
                color: #AEAEB2;
                font-family: 'PT Root UI', sans-serif;
                font-size: 12px;
                font-weight: 600;
                letter-spacing: 1px;
            }
            QPushButton:hover { background: #3A3A3C; color: #FFFFFF; }
        """)
        cancel_btn.clicked.connect(self.reject)

        add_btn = QPushButton("ADD")
        add_btn.setFixedHeight(38)
        add_btn.setCursor(Qt.PointingHandCursor)
        add_btn.setStyleSheet("""
            QPushButton {
                background: #0A84FF;
                border: none;
                border-radius: 10px;
                color: #FFFFFF;
                font-family: 'PT Root UI', sans-serif;
                font-size: 12px;
                font-weight: 700;
                letter-spacing: 1px;
            }
            QPushButton:hover { background: #409CFF; }
        """)
        add_btn.clicked.connect(self._submit)

        buttons.addWidget(cancel_btn)
        buttons.addWidget(add_btn)
        layout.addLayout(buttons)

    def showEvent(self, event):
        super().showEvent(event)
        self._email_input.setFocus()
        self._email_input.selectAll()

    def _submit(self):
        email = self._email_input.text().strip()
        if not email:
            self._error_label.setText("Enter an email address.")
            self._error_label.show()
            return

        if "@" not in email or "." not in email.split("@")[-1]:
            self._error_label.setText("Enter a valid email address.")
            self._error_label.show()
            return

        self.recipient_email = email
        self.accept()


class EmailSenderPage(QWidget):
    @staticmethod
    def _icon_button_stylesheet() -> str:
        return """
            TransparentToolButton, TransparentPushButton {
                background: rgba(28, 28, 30, 0.5);
                border: 1px solid #3A3A3C;
                border-radius: 8px;
                padding: 0;
                color: #8E8E93;
            }
            TransparentToolButton:hover, TransparentPushButton:hover {
                background: rgba(44, 44, 46, 0.85);
                border: 1px solid #0A84FF;
                color: #FFFFFF;
            }
            TransparentToolButton:pressed, TransparentPushButton:pressed {
                background: rgba(58, 58, 60, 0.95);
            }
            TransparentToolButton::menu-indicator {
                image: none;
                width: 0px;
            }
            TransparentPushButton::menu-indicator {
                image: none;
                width: 0px;
            }
        """

    @staticmethod
    def _build_ssl_context() -> ssl.SSLContext:
        """
        Build a deterministic TLS context for SMTP.
        Prefer certifi's CA bundle when available to avoid broken local CA stores
        on some Windows installs, while still keeping certificate verification on.
        """
        cafile = None
        try:
            import certifi
            cafile = certifi.where()
        except Exception:
            cafile = None

        context = ssl.create_default_context(cafile=cafile)
        context.check_hostname = True
        context.verify_mode = ssl.CERT_REQUIRED
        return context

    def __init__(self, parent=None):
        super().__init__(parent)
        self._attachments: list[str] = []
        self._sender_profiles: dict[str, dict[str, str]] = {}
        self._successful_history: dict[str, set[str]] = self._load_outreach_history()
        self._sending = False
        self._stop_requested = False
        self._active_server = None
        self._signals = _SendSignals()
        self._isPreview = False
        self._error_vault: dict[str, str] = {} # recipient -> full error
        self._language = get_language(config_manager.settings.app_language)
        
        self._signals.log.connect(self._on_log)
        self._signals.progress.connect(self._on_progress)
        self._signals.finished.connect(self._on_finished)
        self._signals.error.connect(self._on_error)
        
        self._is_restoring = True
        self._init_widgets()
        self._build_ui()
        self._restore_fields()
        self._connect_buttons()
        self._is_restoring = False

    def _outreach_history_path(self) -> Path:
        return Path(os.getenv("APPDATA", "")) / "ZUGZWANG" / "data" / "outreach_history.txt"

    def _load_outreach_history(self) -> dict[str, set[str]]:
        # Persist across application restarts
        try:
            path = self._outreach_history_path()
            if not path.exists():
                return {}

            history: dict[str, set[str]] = {}
            with open(path, "r", encoding="utf-8") as f:
                for raw_line in f:
                    line = raw_line.strip()
                    if not line or "\t" not in line:
                        continue
                    email, signature = line.split("\t", 1)
                    email = email.strip().lower()
                    signature = signature.strip()
                    if not email or not signature:
                        continue
                    history.setdefault(email, set()).add(signature)
            return history
        except Exception:
            return {}

    def _append_outreach_history(self, email: str, signature: str):
        try:
            path = self._outreach_history_path()
            path.parent.mkdir(parents=True, exist_ok=True)
            with open(path, "a", encoding="utf-8") as f:
                f.write(f"{email.strip().lower()}\t{signature}\n")
        except Exception:
            pass

    def _clear_outreach_history(self):
        self._successful_history = {}
        try:
            path = self._outreach_history_path()
            if path.exists():
                path.unlink()
        except Exception as e:
            self._on_log(f"Could not clear sent history: {e}", "ERROR")
            return

        self._refresh_recipient_history_state()
        self._on_log("Cleared sent email history.", "SUCCESS")

    def _message_signature(self) -> str:
        attachment_parts = []
        for file_path in self._attachments:
            clean_path = str(file_path).strip()
            if not clean_path:
                continue

            size = ""
            mtime = ""
            if os.path.exists(clean_path):
                try:
                    stat = os.stat(clean_path)
                    size = str(stat.st_size)
                    mtime = str(int(stat.st_mtime))
                except OSError:
                    pass
            attachment_parts.append(f"{clean_path}|{size}|{mtime}")

        payload = "\n".join([
            self._smtp_user.text().strip().lower(),
            self._from_name.text().strip(),
            self._reply_to.text().strip().lower(),
            self._subject.text(),
            "html" if self._type_toggle.isChecked() else "plain",
            self._body_text.toPlainText(),
            "\n".join(sorted(attachment_parts)),
        ])
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    def _current_sent_emails(self) -> list[str]:
        signature = self._message_signature()
        return sorted(
            email for email, signatures in self._successful_history.items()
            if signature in signatures
        )

    def _was_message_sent(self, email: str, signature: str | None = None) -> bool:
        email_key = email.strip().lower()
        if not email_key:
            return False
        if signature is None:
            signature = self._message_signature()
        return signature in self._successful_history.get(email_key, set())

    def _record_successful_send(self, email: str, signature: str):
        email_key = email.strip().lower()
        signatures = self._successful_history.setdefault(email_key, set())
        
        # Mark as sent in the database so Edit Page and other components know
        import sqlite3
        from datetime import datetime
        from ..core.config import get_memory_db_path
        from ..core.events import event_bus, EventBus
        try:
            now = datetime.utcnow().isoformat()
            conn = sqlite3.connect(str(get_memory_db_path()), timeout=10.0)
            row = conn.execute("SELECT id FROM leads WHERE email = ? LIMIT 1", (email_key,)).fetchone()
            if row:
                lead_id = row[0]
                conn.execute(
                    "INSERT INTO letter_state (lead_id, sent_at) VALUES (?, ?) "
                    "ON CONFLICT(lead_id) DO UPDATE SET sent_at=excluded.sent_at",
                    (lead_id, now)
                )
                conn.commit()
                event_bus.emit(EventBus.DB_UPDATED, records=[])
            conn.close()
        except Exception as e:
            self._log(f"Failed to update lead status in DB: {e}", "WARNING")
            
        if signature in signatures:
            return
        signatures.add(signature)
        self._append_outreach_history(email_key, signature)

    def _refresh_recipient_history_state(self):
        self._recipient_list.viewport().update()

    def import_emails(self, emails: list[str]):
        """Consolidated API to push leads into the recipient queue."""
        if not emails:
            return
        
        # Merge with existing
        clean_new = [e.strip() for e in emails if e.strip()]
        current = self._get_recipients()
        combined = list(dict.fromkeys(current + clean_new))
        
        self._set_recipients(combined)
        self._on_log(f"Imported {len(clean_new)} lead(s). Total: {len(combined)}", "SUCCESS")

    def _init_widgets(self):
        # Step 1 Widgets
        self._smtp_host = LineEdit(); self._smtp_host.setPlaceholderText(tr("send.placeholder.smtp", self._language))
        self._smtp_port = LineEdit(); self._smtp_port.setText("587"); self._smtp_port.setFixedWidth(70)
        self._smtp_user = LineEdit(); self._smtp_user.setPlaceholderText(tr("send.placeholder.user", self._language))
        self._smtp_pass = LineEdit(); self._smtp_pass.setPlaceholderText(tr("send.placeholder.pass", self._language)); self._smtp_pass.setEchoMode(QLineEdit.Password)
        
        self._auth_switch = MacSwitch()
        self._ssl_switch = MacSwitch()
        self._tls_switch = MacSwitch()
        
        # Step 2 Widgets
        self._from_name = LineEdit(); self._from_name.setPlaceholderText("John Doe")
        self._reply_to = LineEdit(); self._reply_to.setPlaceholderText("reply@domain.com")
        self._subject = LineEdit(); self._subject.setPlaceholderText(tr("send.placeholder.subject", self._language))
        
        self._body_stack = QStackedWidget()
        self._body_stack.setMinimumHeight(80)
        self._body_text = PlainTextEdit()
        self._body_text.setPlaceholderText(tr("send.placeholder.body", self._language))
        self._body_stack.addWidget(self._body_text)
        
        self._body_preview = QTextBrowser()
        self._body_preview.setOpenExternalLinks(True)
        self._body_preview.setOpenLinks(True)
        self._body_stack.addWidget(self._body_preview)
        
        self._type_toggle = MacSwitch()
        self._preview_btn = ToolButton(FluentIcon.VIEW)
        self._attach_btn = ToolButton(FluentIcon.ADD)
        self._attach_clear_btn = ToolButton(FluentIcon.DELETE)
        self._attach_badge = QLabel(tr("send.badge.files", self._language).format(count=0))
        self._attachments = []
        
        # Step 3 Widgets
        self._btn_dedup = PushButton(tr("send.button.clean", self._language))
        self._btn_dedup.setStyleSheet(Theme.secondary_button())
        self._rec_count = QLabel(tr("send.badge.emails", self._language).format(count=0))
        
        self._recipient_list = RecipientListWidget()
        self._recipient_list.setMinimumHeight(0)
        self._recipient_list.items_changed.connect(self._on_recipients_changed)
        
        self._status_log = QTextBrowser()
        self._status_log.setOpenExternalLinks(False)
        self._status_log.setReadOnly(True)
        self._status_log.setText(tr("send.log.initial", self._language))
        
        self._interval_input = LineEdit(); self._interval_input.setText("30")
        self._interval_input.setPlaceholderText("Rec: 30")
        self._interval_input.setFixedWidth(70)
        
        self._batch_size = LineEdit(); self._batch_size.setText("100")
        self._batch_size.setFixedWidth(70)
        
        self._btn_purge_sent = PushButton(tr("send.button.purge", self._language))
        
        self._btn_send_one = PushButton(tr("send.button.send_one", self._language))
        self._btn_stop = PushButton(tr("send.button.stop", self._language))
        self._btn_stop.setEnabled(False)
        self._btn_export = PushButton(tr("send.button.export", self._language))
        self._btn_send_all = PushButton(tr("send.button.send_all", self._language))
        self._btn_clear_data = PushButton(tr("send.button.clear", self._language))
        
        # New Search & Delete
        self._search_input = LineEdit()
        self._search_input.setPlaceholderText(tr("send.placeholder.search", self._language))
        self._search_input.setClearButtonEnabled(True)
        self._search_input.setFixedWidth(150)
        

        
        # Simple Delete Button (Neutral Square)
        self._btn_delete_menu = TransparentToolButton(FluentIcon.DELETE)
        self._btn_delete_menu.setFixedSize(32, 32)
        self._btn_delete_menu.setIconSize(QSize(16, 16))
        self._btn_delete_menu.setCursor(Qt.PointingHandCursor)
        self._btn_delete_menu.setToolTip("Remove recipients")
        self._btn_delete_menu.setStyleSheet(self._icon_button_stylesheet())

        self._sender_completer = QCompleter([], self)
        self._sender_completer.setCaseSensitivity(Qt.CaseInsensitive)
        self._sender_completer.setFilterMode(Qt.MatchContains)
        self._sender_completer.setCompletionMode(QCompleter.PopupCompletion)
        self._smtp_user.setCompleter(self._sender_completer)

    def _build_ui(self):
        self.setObjectName("emailPage")
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        host = QWidget()
        host.setObjectName("emailPageHost")
        host.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        host.setStyleSheet("QWidget#emailPageHost { background: #1C1C1E; }")
        root.addWidget(host)
        
        body = QVBoxLayout(host)
        body.setContentsMargins(24, 16, 24, 16)
        body.setSpacing(10)

        # ── Header Section ───────────────────────────────────────────────────
        header_widget = QWidget()
        header_widget.setFixedHeight(52)
        header_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        header_h = QHBoxLayout(header_widget)
        header_h.setContentsMargins(0, 0, 0, 0)
        header_h.setSpacing(12)
        
        self._page_title = QLabel(tr("send.title", self._language))
        self._page_title.setStyleSheet("color: white; font-family: 'PT Root UI', sans-serif; font-weight: 600; font-size: 28px;")
        header_h.addWidget(self._page_title)
        
        self._status_badge = QLabel(tr("send.status.ready", self._language))
        self._status_badge.setStyleSheet("color: #30D158; font-family: 'PT Root UI', sans-serif; font-weight: 500; font-size: 10px; letter-spacing: 1.4px; border-radius: 6px; padding: 3px 10px;")
        header_h.addWidget(self._status_badge, 0, Qt.AlignVCenter)
        
        header_h.addStretch(1)

        self._btn_send_one.setFixedHeight(36)
        self._btn_send_one.setCursor(Qt.PointingHandCursor)
        self._btn_send_one.setStyleSheet(Theme.zugzwang_success_button())
        header_h.addWidget(self._btn_send_one)
        
        self._btn_stop.setFixedHeight(36)
        self._btn_stop.setCursor(Qt.PointingHandCursor)
        self._btn_stop.setStyleSheet(Theme.secondary_button())
        header_h.addWidget(self._btn_stop)

        self._btn_clear_data.setFixedHeight(36)
        self._btn_clear_data.setStyleSheet(Theme.zugzwang_danger_button())
        self._btn_clear_data.clicked.connect(self._clear_all_data)
        header_h.addWidget(self._btn_clear_data)

        self._btn_send_all.setFixedHeight(36)
        self._btn_send_all.setCursor(Qt.PointingHandCursor)
        self._btn_send_all.setStyleSheet(Theme.zugzwang_primary_button())
        header_h.addWidget(self._btn_send_all)

        body.addWidget(header_widget, 0)
        

        # ── Two-column layout ─────────────────────────────────────────────────
        content_row_widget = QWidget()
        content_row_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        content_row = QHBoxLayout(content_row_widget)
        content_row.setSpacing(12)
        content_row.setContentsMargins(0, 0, 0, 0)

        # Left column (40%): Section 1 + Section 2 stacked
        left_widget = QWidget()
        left_widget.setMinimumWidth(0)
        left_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        left_widget.setStyleSheet("background: transparent; border: none;")
        left_col = QVBoxLayout(left_widget)
        left_col.setSpacing(12)
        left_col.setContentsMargins(0, 0, 0, 0)

        self._card1 = self._build_identity_section()
        self._card1.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        left_col.addWidget(self._card1, 0)

        self._card2 = self._step_card("2", tr("send.step2.title", self._language), self._build_payload_section())
        self._card2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._card2.setMinimumHeight(0)
        left_col.addWidget(self._card2, 1)

        content_row.addWidget(left_widget, 2)

        # Right column (60%): Section 3 full height
        self._card3 = self._step_card("3", tr("send.step3.title", self._language), self._build_monitor_section())
        self._card3.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._card3.setMinimumWidth(0)
        self._card3.setMinimumHeight(0)
        content_row.addWidget(self._card3, 3)

        body.addWidget(content_row_widget, 1)





    def _step_card(self, num: str, title: str, content: QWidget, compact: bool = False) -> QFrame:
        card = QFrame()
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        card.setMinimumHeight(0)
        card.setStyleSheet("QFrame { background: #2C2C2E; border-radius: 14px; border: none; }")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(6)

        hdr = QHBoxLayout()
        hdr.setSpacing(12)
        badge = QLabel(num)
        badge.setFixedSize(20, 20)
        badge.setAlignment(Qt.AlignCenter)
        badge.setStyleSheet("color: white; background: #0A84FF; border-radius: 10px; font-family: 'PT Root UI', sans-serif; font-weight: 700; font-size: 11px;")
        hdr.addWidget(badge)

        ttl = QLabel(title.upper())
        ttl.setStyleSheet("color: #8E8E93; font-family: 'PT Root UI', sans-serif; font-weight: 600; font-size: 11px; letter-spacing: 1.6px; background: transparent; border: none;")
        hdr.addWidget(ttl)
        hdr.addStretch(1)
        layout.addLayout(hdr)

        div = QFrame()
        div.setFixedHeight(1)
        div.setStyleSheet("background: #3A3A3C; border: none;")
        layout.addWidget(div)

        layout.addWidget(content, 1 if not compact else 0)
        return card

    def _build_identity_section(self) -> QWidget:
        from ..core.config import config_manager
        import logging
        s = config_manager.settings

        container = QFrame()
        container.setStyleSheet("QFrame { background: #2C2C2E; border-radius: 14px; border: none; }")
        layout = QVBoxLayout(container)
        layout.setContentsMargins(16, 14, 16, 12)
        layout.setSpacing(8)

        # 1. Header row
        header = QHBoxLayout()
        header.setSpacing(10)
        badge = QLabel("1")
        badge.setFixedSize(20, 20)
        badge.setAlignment(Qt.AlignCenter)
        badge.setStyleSheet("color: white; background: #0A84FF; border-radius: 10px; font-family: '.SF Pro Text', 'PT Root UI', sans-serif; font-weight: 700; font-size: 11px;")
        header.addWidget(badge)

        ttl = QLabel("ZUGZWANG IDENTITY")
        ttl.setStyleSheet("color: #8E8E93; font-family: '.SF Pro Text', 'PT Root UI', sans-serif; font-weight: 600; font-size: 11px; letter-spacing: 1.6px; background: transparent; border: none;")
        header.addWidget(ttl)
        header.addStretch(1)

        settings_btn = QPushButton("⚙ Settings")
        settings_btn.setCursor(Qt.PointingHandCursor)
        settings_btn.setStyleSheet("QPushButton { color: #636366; font-family: '.SF Pro Text', 'PT Root UI', sans-serif; font-size: 12px; background: transparent; border: none; } QPushButton:hover { color: #AEAEB2; }")
        settings_btn.clicked.connect(lambda: self.window()._switch(5))
        header.addWidget(settings_btn)
        layout.addLayout(header)

        # 2. Hairline divider
        div = QFrame()
        div.setFixedHeight(1)
        div.setStyleSheet("background: #3A3A3C; border: none;")
        layout.addSpacing(8)
        layout.addWidget(div)
        layout.addSpacing(8)

        # 3. SMTP status row
        status_row = QHBoxLayout()
        status_row.setSpacing(6)

        status_dot = QLabel("●")
        is_configured = bool(s.email_smtp_user)
        status_color = "#28B84E" if is_configured else "#636366"
        status_dot.setStyleSheet(f"color: {status_color}; font-size: 14px; margin-bottom: 2px; background: transparent; border: none;")
        status_row.addWidget(status_dot)

        srv_label = QLabel(f"{s.email_smtp_host}:{s.email_smtp_port}".upper())
        srv_label.setMinimumWidth(200)
        srv_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        srv_label.setStyleSheet("color: #8E8E93; font-family: ui-monospace, SFMono-Regular, Consolas, monospace; font-size: 12px; background: transparent; border: none;")
        status_row.addWidget(srv_label)
        status_row.addStretch()

        txt_status = QLabel("Connected" if is_configured else "Not configured")
        txt_status.setStyleSheet(f"color: {'#28B84E' if is_configured else '#FF453A'}; font-family: '.SF Pro Text', 'PT Root UI', sans-serif; font-size: 12px; background: transparent; border: none;")
        status_row.addWidget(txt_status)

        test_btn = QPushButton("Test Connection")
        test_btn.setCursor(Qt.PointingHandCursor)
        test_btn.setStyleSheet("QPushButton { color: #636366; font-family: '.SF Pro Text', 'PT Root UI', sans-serif; font-size: 11px; background: transparent; border: none; margin-left: 12px; } QPushButton:hover { color: #0A84FF; }")
        test_btn.clicked.connect(self._test_connection)
        status_row.addWidget(test_btn)

        # Embed status row directly (no wrapper widget)
        layout.addLayout(status_row)
        layout.addSpacing(8)

        # 4. Two-column input grid
        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(4)

        lbl_style = "color: #8E8E93; font-family: '.SF Pro Text', 'PT Root UI', sans-serif; font-size: 10px; font-weight: 600; letter-spacing: 1.3px; text-transform: uppercase; margin-bottom: -2px; background: transparent; border: none;"
        inp_style = "QLineEdit { background: #1C1C1E; border: 1px solid #3A3A3C; border-radius: 8px; color: #F2F2F7; font-family: '.SF Pro Text', 'PT Root UI', sans-serif; font-size: 13px; padding: 0 12px; } QLineEdit:focus { border: 1px solid #0A84FF; }"

        lbl_usr = QLabel("SENDER IDENTITY")
        lbl_usr.setStyleSheet(lbl_style)
        self._smtp_user.setFixedHeight(36)
        self._smtp_user.setStyleSheet(inp_style)
        grid.addWidget(lbl_usr, 0, 0)
        grid.addWidget(self._smtp_user, 1, 0)

        lbl_pass = QLabel("APP PASSWORD")
        lbl_pass.setStyleSheet(lbl_style)
        self._smtp_pass.setFixedHeight(36)
        self._smtp_pass.setStyleSheet(inp_style)
        self._smtp_pass.setEchoMode(QLineEdit.Password)
        grid.addWidget(lbl_pass, 0, 1)
        grid.addWidget(self._smtp_pass, 1, 1)

        # Optional: reconnect save events if disconnected in previous step
        try: self._smtp_user.textChanged.disconnect()
        except: pass
        try: self._smtp_pass.textChanged.disconnect()
        except: pass
        self._smtp_user.textChanged.connect(self._on_sender_identity_changed)
        self._smtp_user.textChanged.connect(self._save_fields)
        self._smtp_pass.textChanged.connect(self._save_fields)

        layout.addLayout(grid)

        # 5. SMTP detail strip
        det_row = QHBoxLayout()
        det_row.setSpacing(16)
        det_row.setContentsMargins(0, 8, 0, 4)

        det_widget = QWidget()
        det_widget.setStyleSheet("background: transparent; border: none;")
        det_widget.setLayout(det_row)

        def add_det(lbl, val, val_color="#8E8E93"):
            l = QLabel(lbl)
            l.setStyleSheet("color: #48484A; font-family: '.SF Pro Text', 'PT Root UI', sans-serif; font-size: 9px; font-weight: 600; letter-spacing: 1.2px; text-transform: uppercase; background: transparent; border: none;")
            v = QLabel(val)
            v.setStyleSheet(f"color: {val_color}; font-family: ui-monospace, SFMono-Regular, Consolas, monospace; font-size: 11px; background: transparent; border: none;")
            h = QHBoxLayout()
            h.setSpacing(5)
            h.addWidget(l)
            h.addWidget(v)
            det_row.addLayout(h)

        def add_sep():
            sep = QFrame()
            sep.setFixedSize(1, 12)
            sep.setStyleSheet("background: #3A3A3C; border: none;")
            det_row.addWidget(sep)

        add_det("SMTP", s.email_smtp_host)
        add_sep()
        add_det("PORT", str(s.email_smtp_port))
        add_sep()
        add_det("TLS", "Enabled", "#28B84E")
        add_sep()
        add_det("AUTH", "Activated", "#28B84E")
        det_row.addStretch()

        layout.addWidget(det_widget)

        return container

    def _build_payload_section(self) -> QWidget:
        container = QWidget()
        container.setMinimumHeight(0)
        container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        container.setStyleSheet("QWidget { background: transparent; border: none; }")
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)
        
        send_widget = QWidget()
        send_widget.setFixedHeight(54)
        send_row = QHBoxLayout(send_widget)
        send_row.setContentsMargins(0, 0, 0, 0)
        send_row.setSpacing(10)
        self._style_input(self._from_name)
        self._style_input(self._reply_to)
        send_row.addLayout(self._make_field(tr("send.field.from_name", self._language), self._from_name), 1)
        send_row.addLayout(self._make_field(tr("send.field.reply_to", self._language), self._reply_to), 1)
        layout.addWidget(send_widget)

        self._style_input(self._subject)
        sub_container = QWidget()
        sub_container.setFixedHeight(54)
        sub_layout = QVBoxLayout(sub_container)
        sub_layout.setContentsMargins(0, 0, 0, 0)
        sub_layout.setSpacing(4)
        sub_layout.addWidget(self._field_label(tr("send.field.subject", self._language)))
        sub_layout.addWidget(self._subject)
        layout.addWidget(sub_container)

        col_body = QVBoxLayout()
        col_body.setContentsMargins(0, 0, 0, 0)
        col_body.setSpacing(4)

        lbl_row_host = QWidget()
        lbl_row_host.setFixedHeight(24)
        lbl_row = QHBoxLayout(lbl_row_host)
        lbl_row.setContentsMargins(0, 0, 0, 0)
        lbl_row.setSpacing(8)
        lbl_row.addWidget(self._field_label("MESSAGE CONTENT"))
        lbl_row.addStretch()
        
        from PySide6.QtCore import QSize
        for btn, tip in [(self._preview_btn, "Preview"), (self._attach_btn, "Attach Files"), (self._attach_clear_btn, "Clear Attachments")]:
            btn.setFixedSize(24, 24)
            btn.setIconSize(QSize(16, 16))
            btn.setCursor(Qt.PointingHandCursor)
            btn.setToolTip(tip)
            btn.setStyleSheet("""
                ToolButton { background: transparent; border: none; border-radius: 6px; color: #636366; }
                ToolButton:hover { color: #FFFFFF; }
            """)
            lbl_row.addWidget(btn)
        
        self._attach_clear_btn.setStyleSheet("""
            ToolButton { background: transparent; border: none; border-radius: 6px; color: #636366; }
            ToolButton:hover { color: #FF453A; }
        """)
        
        self._attach_badge.setStyleSheet("color: #636366; font-family: 'PT Root UI', monospace; font-size: 11px; background: transparent;")
        lbl_row.addWidget(self._attach_badge)
        
        lbl_row.addSpacing(14)
        
        html_lbl = QLabel("HTML")
        html_lbl.setStyleSheet("color: #8E8E93; font-family: 'PT Root UI', sans-serif; font-size: 12px; background: transparent; border: none;")
        lbl_row.addWidget(html_lbl)
        
        self._type_toggle.toggled.connect(self._save_fields)
        lbl_row.addWidget(self._type_toggle)
        col_body.addWidget(lbl_row_host)
 
        self._body_stack.setMinimumHeight(80)
        self._body_stack.setMinimumWidth(0)
        self._body_stack.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._style_plaintext(self._body_text)
        col_body.addWidget(self._body_stack, 1)
        layout.addLayout(col_body, 1)
        return container

    def _build_monitor_section(self) -> QWidget:
        container = QWidget()
        container.setMinimumHeight(0)
        container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        container.setStyleSheet("QWidget { background: transparent; border: none; }")
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 5, 0, 0)
        layout.setSpacing(10)

        rec_widget = QWidget()
        rec_widget.setFixedHeight(44)
        rec_h = QHBoxLayout(rec_widget)
        rec_h.setContentsMargins(0, 0, 0, 0)
        rec_h.setSpacing(12)
        
        self._style_input(self._interval_input)
        self._interval_input.setFixedHeight(34)
        self._interval_input.setFixedWidth(52)
        
        iv_lbl = self._field_label("INTERVAL (S)")
        rec_h.addWidget(iv_lbl, 0, Qt.AlignVCenter)
        rec_h.addWidget(self._interval_input, 0, Qt.AlignVCenter)
        
        bs_lbl = self._field_label(tr("send.field.batch", self._language))
        rec_h.addWidget(bs_lbl, 0, Qt.AlignVCenter)
        
        self._style_input(self._batch_size)
        self._batch_size.setFixedHeight(34)
        self._batch_size.setFixedWidth(64)
        rec_h.addWidget(self._batch_size, 0, Qt.AlignVCenter)
        
        rec_h.addStretch(1)

        self._btn_dedup.setFixedHeight(34)
        self._btn_dedup.setFixedWidth(100)
        self._btn_dedup.setStyleSheet(Theme.zugzwang_danger_button())
        rec_h.addWidget(self._btn_dedup, 0, Qt.AlignVCenter)
        
        self._btn_purge_sent.setFixedHeight(34)
        self._btn_purge_sent.setFixedWidth(140)
        self._btn_purge_sent.setStyleSheet(Theme.zugzwang_danger_button())
        rec_h.addWidget(self._btn_purge_sent, 0, Qt.AlignVCenter)
        
        self._rec_count.setStyleSheet("color: #0A84FF; font-family: 'PT Root UI', monospace; font-size: 11px; font-weight: 600; background: transparent;")
        # rec_h.addWidget(self._rec_count, 0, Qt.AlignVCenter) # Moved down
        layout.addWidget(rec_widget)

        # Side-by-Side Monitor Area (Grid for perfect 50/50 split)
        monitor_split = QGridLayout()
        monitor_split.setSpacing(12)
        monitor_split.setColumnStretch(0, 60)
        monitor_split.setColumnStretch(1, 40)
        
        # Left: Recipients
        rec_col = QVBoxLayout()
        rec_col.setContentsMargins(0, 0, 0, 0)
        rec_col.setSpacing(12)
        
        lbl_rec_row = QHBoxLayout()
        lbl_rec_row.setSpacing(12)
        lbl_rec = self._field_label("QUEUE")
        lbl_rec.setFixedHeight(32)
        lbl_rec_row.addWidget(lbl_rec)
        
        self._rec_count.setStyleSheet("color: #0A84FF; font-family: 'PT Root UI', sans-serif; font-size: 11px; font-weight: 600; background: transparent;")
        lbl_rec_row.addWidget(self._rec_count)
        lbl_rec_row.addStretch(1)
        
        # Add Search Input
        self._search_input.setMaximumWidth(180)
        self._search_input.setMinimumWidth(100)
        self._search_input.setFixedHeight(32)
        self._search_input.setPlaceholderText("Find recipient...")
        self._search_input.setStyleSheet("""
            QLineEdit {
                background: rgba(28, 28, 30, 0.5);
                border: 1px solid #3A3A3C;
                border-radius: 8px;
                color: #FFFFFF;
                font-family: 'PT Root UI', sans-serif;
                font-size: 13px;
                padding: 0 10px;
            }
            QLineEdit:focus {
                background: rgba(44, 44, 46, 0.7);
                border: 1px solid #0A84FF;
            }
        """)
        self._search_input.textChanged.connect(self._on_search_recipients)
        self._status_log.anchorClicked.connect(self._on_error_anchor_clicked)
        lbl_rec_row.addWidget(self._search_input, 0, Qt.AlignVCenter)
        
        self._setup_delete_menu()

        self._btn_add_manual_recipient = TransparentToolButton(FluentIcon.EDIT)
        self._btn_add_manual_recipient.setFixedSize(32, 32)
        self._btn_add_manual_recipient.setIconSize(QSize(16, 16))
        self._btn_add_manual_recipient.setToolTip("Add recipient manually")
        self._btn_add_manual_recipient.setCursor(Qt.PointingHandCursor)
        self._btn_add_manual_recipient.setStyleSheet(self._icon_button_stylesheet())
        self._btn_add_manual_recipient.clicked.connect(self._on_add_manual_recipient)
        
        # Import Button (shows a popup menu with all import sources)
        self._btn_load_leads = TransparentToolButton(FluentIcon.ADD)
        self._btn_load_leads.setFixedSize(32, 32)
        self._btn_load_leads.setIconSize(QSize(16, 16))
        self._btn_load_leads.setToolTip("Import recipients…")
        self._btn_load_leads.setCursor(Qt.PointingHandCursor)
        self._btn_load_leads.setStyleSheet(self._icon_button_stylesheet())
        self._btn_load_leads.clicked.connect(self._show_import_menu)
        
        lbl_rec_row.addWidget(self._btn_load_leads, 0, Qt.AlignVCenter)
        lbl_rec_row.addWidget(self._btn_add_manual_recipient, 0, Qt.AlignVCenter)
        lbl_rec_row.addWidget(self._btn_delete_menu, 0, Qt.AlignVCenter)
        
        rec_col.addLayout(lbl_rec_row)
        
        # Re-use existing recipients list instead of plaintext
        # ── Improvement 7: Empty State Guidance ──────────────────────────────
        from PySide6.QtWidgets import QStackedWidget
        self._rec_stack = QStackedWidget()
        
        from .components import EmptyStateWidget
        self._rec_empty = EmptyStateWidget(
            FluentIcon.SEARCH,
            title="Queue is Empty",
            description="Import emails, paste from clipboard, or load directly from your lead database.",
            button_text="LOAD LEADS",
            button_callback=self._load_from_leads
        )
        
        self._rec_stack.addWidget(self._rec_empty)
        self._rec_stack.addWidget(self._recipient_list)
        self._rec_stack.setCurrentWidget(self._rec_empty)
        
        rec_col.addWidget(self._rec_stack, 1)
        monitor_split.addLayout(rec_col, 0, 0)
        
        # Right: Log Stack
        log_stack = QVBoxLayout()
        log_stack.setContentsMargins(0, 0, 0, 0)
        log_stack.setSpacing(12)
        
        log_header = QHBoxLayout()
        log_header.setSpacing(12)
        lbl_log = self._field_label("LOG")
        lbl_log.setFixedHeight(32) # Perfectly vertically aligned with left side
        log_header.addWidget(lbl_log)

        log_header.addStretch(1)


        
        # Copy Log Button
        self._btn_copy_log = TransparentToolButton(FluentIcon.COPY)
        self._btn_copy_log.setFixedSize(32, 32)
        self._btn_copy_log.setIconSize(QSize(16, 16))
        self._btn_copy_log.setToolTip("Copy Activity Log")
        self._btn_copy_log.setCursor(Qt.PointingHandCursor)
        self._btn_copy_log.setStyleSheet(self._icon_button_stylesheet())
        self._btn_copy_log.clicked.connect(lambda: QGuiApplication.clipboard().setText(self._status_log.toPlainText()))
        log_header.addWidget(self._btn_copy_log, 0, Qt.AlignVCenter)

        self._btn_clear_history = TransparentToolButton(FluentIcon.BROOM)
        self._btn_clear_history.setFixedSize(32, 32)
        self._btn_clear_history.setIconSize(QSize(16, 16))
        self._btn_clear_history.setToolTip("Clear Sent History")
        self._btn_clear_history.setCursor(Qt.PointingHandCursor)
        self._btn_clear_history.setStyleSheet(self._icon_button_stylesheet())
        log_header.addWidget(self._btn_clear_history, 0, Qt.AlignVCenter)

        # Clear Log Button
        self._btn_clear_log = TransparentToolButton(FluentIcon.DELETE)
        self._btn_clear_log.setFixedSize(32, 32)
        self._btn_clear_log.setIconSize(QSize(16, 16))
        self._btn_clear_log.setToolTip("Clear Activity Log")
        self._btn_clear_log.setCursor(Qt.PointingHandCursor)
        self._btn_clear_log.setStyleSheet(self._icon_button_stylesheet())
        self._btn_clear_log.clicked.connect(lambda: self._status_log.clear())
        log_header.addWidget(self._btn_clear_log, 0, Qt.AlignVCenter)
        log_stack.addLayout(log_header)
        
        self._status_log.setStyleSheet("""
            QTextBrowser, QTextEdit {
                background: #1A1A1A;
                border: 1px solid #3A3A3C;
                border-radius: 10px;
                padding: 12px;
                color: #8E8E93;
                font-family: 'PT Root UI', monospace;
                font-size: 12px;
                line-height: 1.6;
            }
        """)
        self._status_log.setLineWrapMode(QTextBrowser.WidgetWidth)
        self._status_log.setMinimumWidth(0)
        self._status_log.setMinimumHeight(0)
        self._status_log.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # Remove fixed height to let it expand naturally with the recipient panel
        log_stack.addWidget(self._status_log, 1)
        
        monitor_split.addLayout(log_stack, 0, 1)
        layout.addLayout(monitor_split, 1)

        # Gmail Policy Notice - Spanning Full Width at Bottom
        notice_card = QFrame()
        notice_card.setObjectName("NoticeCard")
        notice_card.setStyleSheet("QFrame#NoticeCard { background: rgba(10, 132, 255, 0.08); border: none; border-radius: 10px; }")
        notice_h = QHBoxLayout(notice_card); notice_h.setContentsMargins(12, 10, 12, 10); notice_h.setSpacing(10)
        
        from qfluentwidgets import IconWidget
        n_ic = IconWidget(FluentIcon.INFO); n_ic.setFixedSize(16, 16); n_ic.setStyleSheet("background: transparent; border: none; color: #0A84FF;")
        notice_h.addWidget(n_ic)
        
        n_txt = QLabel(tr("send.notice.gmail", self._language))
        n_txt.setStyleSheet("color: #4EA1FF; font-size: 11px; font-weight: 500; font-family: 'PT Root UI', sans-serif;")
        notice_h.addWidget(n_txt, 1)
        
        layout.addWidget(notice_card)

        return container

    def _clear_all_data(self):
        # Keep SMTP transport settings intact. "Clear all" should reset compose
        # state, not wipe the configured mail server out of app settings.
        self._smtp_user.clear()
        self._smtp_pass.clear()
        self._auth_switch.setChecked(True)
        self._ssl_switch.setChecked(False)
        self._tls_switch.setChecked(True)

        # Payload
        self._from_name.clear()
        self._reply_to.clear()
        self._subject.clear()
        self._body_text.clear()
        self._type_toggle.setChecked(False)
        
        # Persistence
        self._save_fields()

    def _field_label(self, text: str) -> QLabel:
        label = QLabel(text.upper())
        label.setStyleSheet("color: #8E8E93; font-family: 'PT Root UI', sans-serif; font-size: 10px; font-weight: 600; letter-spacing: 1.3px; background: transparent; border: none;")
        return label

    def _make_field(self, label: str, widget: QWidget) -> QVBoxLayout:
        col = QVBoxLayout()
        col.setSpacing(4)
        col.addWidget(self._field_label(label))
        col.addWidget(widget)
        return col

    def _style_input(self, widget: QWidget) -> None:
        widget.setFixedHeight(34)

    def _style_plaintext(self, widget: QWidget) -> None:
        pass

    def _make_switch_row(self, text: str, switch: MacSwitch, on_color: str) -> QHBoxLayout:
        h = QHBoxLayout()
        h.setSpacing(10)
        lbl = QLabel(text)
        lbl.setStyleSheet("color: #AEAEB2; font-family: 'PT Root UI', sans-serif; font-size: 10px; font-weight: 500; background: transparent; border: none;")
        
        switch.setOnColor(on_color)
        
        h.addWidget(switch)
        h.addWidget(lbl)
        h.setAlignment(lbl, Qt.AlignLeft)
        h.addStretch()
        return h


    # ── Signal Handlers & Logic ──────────────────────────────────────────────
    def _connect_buttons(self):
        self._btn_clear_data.clicked.connect(self._clear_all_data)
        self._preview_btn.clicked.connect(self._toggle_preview)
        self._attach_btn.clicked.connect(self._on_attach)
        self._attach_clear_btn.clicked.connect(self._on_attach_clear)
        self._btn_dedup.clicked.connect(self._on_dedup)
        self._btn_send_one.clicked.connect(self._send_test)
        self._btn_send_all.clicked.connect(self._send_all)
        self._btn_stop.clicked.connect(self._on_stop)
        self._btn_purge_sent.clicked.connect(self._on_purge_sent)
        self._btn_clear_history.clicked.connect(self._clear_outreach_history)
        self._btn_export.clicked.connect(self._on_export)

        # self._recipients_text is replaced by self._recipient_list
        # The list widget handles changes via items_changed which is already connected in _init_widgets
        
        # Persistence hooks for all fields

        self._smtp_host.textChanged.connect(self._save_fields)
        self._smtp_port.textChanged.connect(self._save_fields)
        self._smtp_user.textChanged.connect(self._save_fields)
        self._smtp_user.textChanged.connect(self._refresh_recipient_history_state)
        self._smtp_pass.textChanged.connect(self._save_fields)
        self._from_name.textChanged.connect(self._save_fields)
        self._from_name.textChanged.connect(self._refresh_recipient_history_state)
        self._reply_to.textChanged.connect(self._save_fields)
        self._reply_to.textChanged.connect(self._refresh_recipient_history_state)
        self._subject.textChanged.connect(self._save_fields)
        self._subject.textChanged.connect(self._refresh_recipient_history_state)
        self._body_text.textChanged.connect(self._save_fields)
        self._body_text.textChanged.connect(self._refresh_preview_if_open)
        self._body_text.textChanged.connect(self._refresh_recipient_history_state)
        self._interval_input.textChanged.connect(self._save_fields)
        
        self._auth_switch.toggled.connect(self._save_fields)
        self._ssl_switch.toggled.connect(self._save_fields)
        self._tls_switch.toggled.connect(self._save_fields)
        self._type_toggle.toggled.connect(self._save_fields)
        self._type_toggle.toggled.connect(self._refresh_preview_if_open)
        self._type_toggle.toggled.connect(self._refresh_recipient_history_state)
        self._sender_completer.activated[str].connect(self._apply_sender_profile)

    def _on_log(self, message: str, level: str):
        color = "#FFFFFF"
        if level == "WARNING": color = "#FF9F0A"
        elif level == "ERROR": color = "#FF453A"
        elif level == "SUCCESS": color = "#32D74B"
        
        self._status_log.append(f'<span style="color: {color}">{message}</span>')
        self._status_log.verticalScrollBar().setValue(self._status_log.verticalScrollBar().maximum())

    def _on_progress(self, current: int, total: int):
        self._status_badge.setText(f"{tr('send.status.sending', self._language)} {current}/{total}")

    def _on_finished(self, sent: int, failed: int, was_aborted: bool):
        self._set_sending_state(False)
        self._status_badge.setText(tr("monitor.status.cancelled", self._language) if was_aborted else tr("send.status.done", self._language))
        msg = f"Broadcast aborted by user. Sent: {sent}, Failed: {failed}" if was_aborted else f"Broadcast concluded. Sent: {sent}, Failed: {failed}"
        color = "#FF453A" if was_aborted else "#4EA1FF"
        self._status_log.append(f'<span style="color: {color}">{msg}</span>')

    def _on_error(self, message: str):
        self._on_log(message, "ERROR")

    def _on_error_anchor_clicked(self, url):
        recipient = url.toString().split(":")[-1]
        error_msg = self._error_vault.get(recipient, "No detailed error captured.")
        
        from .components import ZugzwangDialog
        box = ZugzwangDialog("Transmission Error", f"Recipient: {recipient}\n\n{error_msg}", self, single_button=True)
        box.exec()

    def _toggle_preview(self):
        self._isPreview = not self._isPreview
        if self._isPreview:
            self._render_preview_content()
            self._body_stack.setCurrentIndex(1)
            self._preview_btn.setText("EDIT")
        else:
            self._body_stack.setCurrentIndex(0)
            self._preview_btn.setText("PREVIEW")

    def _refresh_preview_if_open(self):
        if self._isPreview:
            self._render_preview_content()

    def _render_preview_content(self):
        raw = self._body_text.toPlainText()
        if self._type_toggle.isChecked():
            rendered = self._wrap_email_preview_html(raw)
        else:
            plain_html = html.escape(raw).replace("\n", "<br>")
            rendered = self._wrap_email_preview_html(f"<div class='plain-body'>{plain_html}</div>")
        self._body_preview.setHtml(rendered)

    def _wrap_email_preview_html(self, body_html: str) -> str:
        return f"""
<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <style>
    body {{
      margin: 0;
      padding: 20px;
      background: #1C1C1E;
      color: #E5E5EA;
      font-family: 'PT Root UI', -apple-system, 'Segoe UI', sans-serif;
      line-height: 1.6;
    }}
    .preview-shell {{
      max-width: 760px;
      margin: 0 auto;
      background: #242426;
      border: 1px solid #3A3A3C;
      border-radius: 14px;
      overflow: hidden;
      box-shadow: 0 14px 40px rgba(0, 0, 0, 0.28);
    }}
    .preview-bar {{
      padding: 10px 16px;
      background: #2F2F31;
      border-bottom: 1px solid #3A3A3C;
      color: #8E8E93;
      font-size: 11px;
      font-weight: 600;
      letter-spacing: 1.2px;
    }}
    .preview-body {{
      padding: 24px;
      color: #F2F2F7;
      word-break: break-word;
    }}
    .preview-body a {{
      color: #4EA1FF;
      text-decoration: none;
    }}
    .preview-body img {{
      max-width: 100%;
      height: auto;
    }}
    .preview-body table {{
      max-width: 100%;
      border-collapse: collapse;
    }}
    .plain-body {{
      white-space: normal;
    }}
  </style>
</head>
<body>
  <div class="preview-shell">
    <div class="preview-bar">EMAIL PREVIEW</div>
    <div class="preview-body">{body_html}</div>
  </div>
</body>
</html>
"""

    def _on_attach(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Attach Files")
        if files:
            self._attachments = list(dict.fromkeys(files))
            self._update_attachment_badge()
            self._refresh_recipient_history_state()
            self._save_fields()
            config_manager.flush()
            self._on_log(f"Attached {len(files)} file(s).", "INFO")

    def _on_attach_clear(self):
        if self._attachments:
            count = len(self._attachments)
            self._attachments = []
            self._update_attachment_badge()
            self._refresh_recipient_history_state()
            self._save_fields()
            config_manager.flush()
            self._on_log(f"Cleared {count} attachment(s).", "WARNING")

    def _update_attachment_badge(self):
        count = len(self._attachments)
        self._attach_badge.setText(
            tr("send.badge.files", self._language).format(count=count)
        )
        color = "#32D74B" if count > 0 else "#FF6B6B"
        self._attach_badge.setStyleSheet(
            f"color: {color}; font-family: 'PT Root UI', monospace; font-size: 11px; background: transparent;"
        )

    def _get_recipients(self) -> list[str]:
        return [self._recipient_list.item(i).email for i in range(self._recipient_list.count())]

    def _load_sender_profiles(self):
        raw_profiles = getattr(config_manager.settings, "email_sender_profiles", []) or []
        profiles: dict[str, dict[str, str]] = {}
        for entry in raw_profiles:
            if not isinstance(entry, dict):
                continue
            email = str(entry.get("email", "")).strip()
            if not self._is_valid_sender_profile_email(email):
                continue
            profiles[email.lower()] = {
                "email": email,
                "password": str(entry.get("password", "")),
            }
        self._sender_profiles = profiles
        self._refresh_sender_completer()

    def _refresh_sender_completer(self):
        emails = [entry["email"] for entry in self._sender_profiles.values()]
        emails.sort(key=str.lower)
        self._sender_completer.model().setStringList(emails)

    def _remember_current_sender_profile(self):
        email = self._smtp_user.text().strip()
        password = self._smtp_pass.text()
        if not self._is_valid_sender_profile_email(email) or not password:
            return
        self._sender_profiles[email.lower()] = {
            "email": email,
            "password": password,
        }
        if len(self._sender_profiles) > 20:
            ordered = list(sorted(self._sender_profiles.values(), key=lambda item: item["email"].lower()))
            ordered = ordered[-20:]
            self._sender_profiles = {item["email"].lower(): item for item in ordered}
        self._refresh_sender_completer()

    @staticmethod
    def _is_valid_sender_profile_email(email: str) -> bool:
        return bool(_EMAIL_PROFILE_RE.match((email or "").strip()))

    def _apply_sender_profile(self, email: str):
        profile = self._sender_profiles.get(email.strip().lower())
        if not profile:
            return
        self._is_restoring = True
        self._smtp_user.setText(profile["email"])
        self._smtp_pass.setText(profile.get("password", ""))
        self._is_restoring = False
        self._save_fields()

    def _on_sender_identity_changed(self):
        if getattr(self, "_is_restoring", False):
            return
        email = self._smtp_user.text().strip().lower()
        profile = self._sender_profiles.get(email)
        if profile and self._smtp_pass.text() != profile.get("password", ""):
            self._is_restoring = True
            self._smtp_pass.setText(profile.get("password", ""))
            self._is_restoring = False

    def _set_recipients(self, emails: list[str]):
        self._recipient_list.setUpdatesEnabled(False)
        self._recipient_list.clear()
        for e in emails:
            e = e.strip()
            if e:
                self._recipient_list.addItem(RecipientItem(e))
        self._recipient_list.setUpdatesEnabled(True)
        self._recipient_list.viewport().update()
        self._on_recipients_changed()

    def _on_dedup(self):
        emails = self._get_recipients()
        unique = list(dict.fromkeys(emails))
        if len(unique) < len(emails):
            self._set_recipients(unique)
            self._on_log(f"De-duplicated list. {len(unique)} unique recipients.", "SUCCESS")
        else:
            self._on_log("Queue is already deduplicated.", "INFO")

    def _on_stop(self):
        self._stop_requested = True
        self._on_log("Stop requested by user...", "WARNING")
        # Aggressive Stop: Force close the connection if it exists
        if self._active_server:
            try:
                self._active_server.close()
                self._on_log("Active connection forced to close.", "WARNING")
            except: pass

    def _on_purge_sent(self):
        sent_emails = {email.strip().lower() for email in self._successful_history.keys() if email.strip()}
        if not sent_emails:
            self._on_log("No successful transmissions to purge.", "WARNING")
            return
            
        current = self._get_recipients()
        purged = [e for e in current if e.strip().lower() not in sent_emails]

        if len(purged) == len(current):
            self._on_log("No queued recipients matched sent history.", "WARNING")
            return
        
        self._set_recipients(purged)
        self._on_log(f"Purged sent emails from queue. {len(purged)} remaining.", "SUCCESS")

    def _setup_delete_menu(self):
        self._delete_menu = RoundMenu(parent=self)
        del_sel = Action(FluentIcon.DELETE, "Delete Selection", self)
        del_sel.triggered.connect(self._on_delete_selection)
        
        del_all = Action(FluentIcon.DELETE, "Delete All", self)
        del_all.triggered.connect(self._on_delete_all)
        
        self._delete_menu.addAction(del_sel)
        self._delete_menu.addAction(del_all)
        self._btn_delete_menu.clicked.connect(self._show_delete_menu)

    def _show_delete_menu(self):
        if not hasattr(self, "_delete_menu"):
            return
        pos = self._btn_delete_menu.mapToGlobal(self._btn_delete_menu.rect().bottomLeft())
        self._delete_menu.exec(pos)

    def _on_add_manual_recipient(self):
        dialog = ManualRecipientDialog(self)
        if self.window():
            center = self.window().geometry().center()
            dialog.move(center.x() - dialog.width() // 2, center.y() - dialog.height() // 2)

        if dialog.exec() != QDialog.Accepted:
            return

        email = dialog.recipient_email.strip()
        email_key = email.lower()
        existing = [self._recipient_list.item(i).email.lower() for i in range(self._recipient_list.count())]
        if email_key in existing:
            self._on_log(f"{email} is already in the queue.", "WARNING")
            return

        self._recipient_list.addItem(RecipientItem(email))
        added_item = self._recipient_list.item(self._recipient_list.count() - 1)
        self._recipient_list.setCurrentItem(added_item)
        self._recipient_list.scrollToItem(added_item)
        self._on_recipients_changed()
        self._on_log(f"Added {email} to the queue.", "SUCCESS")

    def _on_delete_all(self):
        self._recipient_list.clear()
        self._on_recipients_changed()
        self._on_log("Cleared all recipients.", "WARNING")

    def _on_delete_selection(self):
        selected = self._recipient_list.selectedItems()
        if not selected:
            self._on_log("No selection to delete.", "WARNING")
            return
        
        count = len(selected)
        for item in selected:
            row = self._recipient_list.row(item)
            self._recipient_list.takeItem(row)
            
        self._on_recipients_changed()
        self._on_log(f"Deleted selection ({count} rows).", "SUCCESS")

    def _on_search_recipients(self, text: str):
        text = text.strip().lower()
        
        # ── Option 4: The Easter Egg ──
        if text == "/coffee":
            from PySide6.QtGui import QDesktopServices
            from PySide6.QtCore import QUrl
            QDesktopServices.openUrl(QUrl("https://wa.me/212663007212?text=slm%20khoya%20khdmt%20b%20app%20dylk%20w%20bghit%20n%20supportik,"))
            self._search_input.clear()
            self._on_log("☕ You found the secret coffee machine! Thanks for the support!", "SUCCESS")
            return
            
        if not text:
            for i in range(self._recipient_list.count()):
                self._recipient_list.item(i).setHidden(False)
            return

        for i in range(self._recipient_list.count()):
            item = self._recipient_list.item(i)
            item.setHidden(text not in item.email.lower())

    def _on_export(self):
        current_sent = self._current_sent_emails()
        if not current_sent:
            self._on_log("No successful transmissions to export.", "WARNING")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export Successful Emails", "broadcast_sent.txt", "Text Files (*.txt)")
        if path:
            with open(path, "w") as f: f.write("\n".join(current_sent))
            self._on_log(f"Exported to {Path(path).name}", "SUCCESS")

    def _show_import_menu(self):
        """Show a popup menu with all import source options."""
        from PySide6.QtWidgets import QMenu
        from PySide6.QtGui import QCursor
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background: #2C2C2E;
                color: #FFFFFF;
                border: 1px solid #3A3A3C;
                border-radius: 10px;
                font-family: 'PT Root UI', sans-serif;
                font-size: 13px;
                padding: 4px;
            }
            QMenu::item {
                padding: 8px 20px 8px 12px;
                border-radius: 6px;
            }
            QMenu::item:selected { background: #0A84FF; }
            QMenu::separator { height: 1px; background: #3A3A3C; margin: 4px 0; }
        """)
        from qfluentwidgets import FluentIcon
        act_leads   = menu.addAction("  Load from Leads DB")
        menu.addSeparator()
        act_pdf     = menu.addAction("  Import from PDF")
        act_docx    = menu.addAction("  Import from Word (.docx)")
        act_excel   = menu.addAction("  Import from Excel (.xlsx)")
        menu.addSeparator()
        act_clip    = menu.addAction("  Paste from Clipboard")

        chosen = menu.exec(QCursor.pos())
        if chosen == act_leads:
            self._load_from_leads()
        elif chosen == act_pdf:
            self._import_from_pdf()
        elif chosen == act_docx:
            self._import_from_docx()
        elif chosen == act_excel:
            self._import_from_excel()
        elif chosen == act_clip:
            self._import_from_clipboard()

    def _extract_emails_from_text(self, text: str) -> list[str]:
        """Extract unique valid email addresses from a block of text."""
        import re
        pattern = r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}'
        found = re.findall(pattern, text)
        return list(dict.fromkeys(e.lower().strip() for e in found))

    def _import_from_pdf(self):
        """Open a PDF file and extract email addresses from its text content."""
        path, _ = QFileDialog.getOpenFileName(self, "Import from PDF", "", "PDF Files (*.pdf)")
        if not path:
            return
        try:
            import importlib.util
            if importlib.util.find_spec("pypdf") is not None:
                from pypdf import PdfReader
                reader = PdfReader(path)
                text = "\n".join(page.extract_text() or "" for page in reader.pages)
            elif importlib.util.find_spec("pdfplumber") is not None:
                import pdfplumber
                with pdfplumber.open(path) as pdf:
                    text = "\n".join(p.extract_text() or "" for p in pdf.pages)
            else:
                self._on_log("PDF import requires 'pypdf' or 'pdfplumber'. Install via pip.", "ERROR")
                return
            emails = self._extract_emails_from_text(text)
            if emails:
                self.import_emails(emails)
                self._on_log(f"PDF import: found {len(emails)} email(s) in '{Path(path).name}'.", "SUCCESS")
            else:
                self._on_log(f"No email addresses found in '{Path(path).name}'.", "WARNING")
        except Exception as e:
            self._on_log(f"PDF import error: {e}", "ERROR")

    def _import_from_docx(self):
        """Open a .docx file and extract email addresses."""
        path, _ = QFileDialog.getOpenFileName(self, "Import from Word Document", "", "Word Documents (*.docx)")
        if not path:
            return
        try:
            import importlib.util
            if importlib.util.find_spec("docx") is None:
                self._on_log("DOCX import requires 'python-docx'. Install via pip.", "ERROR")
                return
            from docx import Document
            doc = Document(path)
            text = "\n".join(para.text for para in doc.paragraphs)
            emails = self._extract_emails_from_text(text)
            if emails:
                self.import_emails(emails)
                self._on_log(f"DOCX import: found {len(emails)} email(s) in '{Path(path).name}'.", "SUCCESS")
            else:
                self._on_log(f"No email addresses found in '{Path(path).name}'.", "WARNING")
        except Exception as e:
            self._on_log(f"DOCX import error: {e}", "ERROR")

    def _import_from_excel(self):
        """Open a .xlsx/.xls/.csv file and extract email addresses."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Import from Spreadsheet", "",
            "Spreadsheet Files (*.xlsx *.xls *.csv);;All Files (*)"
        )
        if not path:
            return
        try:
            suffix = Path(path).suffix.lower()
            text = ""
            if suffix == ".csv":
                import csv
                with open(path, newline="", encoding="utf-8", errors="replace") as f:
                    text = "\n".join(",".join(row) for row in csv.reader(f))
            else:
                import importlib.util
                if importlib.util.find_spec("openpyxl") is not None:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    rows = []
                    for ws in wb.worksheets:
                        for row in ws.iter_rows(values_only=True):
                            rows.append(" ".join(str(c) for c in row if c is not None))
                    text = "\n".join(rows)
                elif importlib.util.find_spec("xlrd") is not None:
                    import xlrd
                    wb = xlrd.open_workbook(path)
                    rows = []
                    for sheet in wb.sheets():
                        for rx in range(sheet.nrows):
                            rows.append(" ".join(str(sheet.cell(rx, cx).value) for cx in range(sheet.ncols)))
                    text = "\n".join(rows)
                else:
                    self._on_log("Excel import requires 'openpyxl'. Install via pip.", "ERROR")
                    return
            emails = self._extract_emails_from_text(text)
            if emails:
                self.import_emails(emails)
                self._on_log(f"Excel import: found {len(emails)} email(s) in '{Path(path).name}'.", "SUCCESS")
            else:
                self._on_log(f"No email addresses found in '{Path(path).name}'.", "WARNING")
        except Exception as e:
            self._on_log(f"Excel import error: {e}", "ERROR")

    def _import_from_clipboard(self):
        """Extract email addresses from the current clipboard text."""
        from PySide6.QtGui import QGuiApplication
        text = QGuiApplication.clipboard().text()
        if not text.strip():
            self._on_log("Clipboard is empty.", "WARNING")
            return
        emails = self._extract_emails_from_text(text)
        if emails:
            self.import_emails(emails)
            self._on_log(f"Clipboard import: found {len(emails)} email(s).", "SUCCESS")
        else:
            self._on_log("No email addresses found in clipboard.", "WARNING")

    def _load_from_leads(self):
        from .load_leads_dialog import LoadLeadsDialog
        from .event_bridge import event_bus
        dialog = LoadLeadsDialog(self)
        if dialog.exec():
            emails = dialog.selected_emails
            if emails:
                self.import_emails(emails)
                event_bus.emit(
                    "toast.show",
                    title="Leads Loaded",
                    subtitle=f"Imported {len(emails)} emails into the broadcast queue.",
                    type="success"
                )

    # import_emails consolidated above

    def _send_test(self):
        recs = self._get_recipients()
        if not recs: return self._on_log("No recipients in queue.", "ERROR")
        
        from ..core.security import LicenseManager
        if not LicenseManager.can_send_email(1):
            from .activation_dialog import ActivationDialog
            status = LicenseManager.get_email_trial_status()
            self._signals.log.emit(f"Send Failed: Reached daily limit of {status['total']} emails.", "ERROR")
            dlg = ActivationDialog(self)
            dlg.exec()
            if not LicenseManager.can_send_email(1):
                return
                
        self._log(f"Executing test broadcast to {recs[0]}...")
        self._set_sending_state(True)
        threading.Thread(target=self._worker_send, args=([recs[0]],), daemon=True).start()

    def _send_all(self):
        recs = self._get_recipients()
        if not recs: return self._on_log("Recipient queue empty.", "ERROR")
        
        from ..core.security import LicenseManager
        count = len(recs)
        status = LicenseManager.get_email_trial_status()
        
        if not LicenseManager.is_active() and count > status['remaining']:
            if status['remaining'] <= 0:
                from .activation_dialog import ActivationDialog
                self._signals.log.emit(f"Broadcast blocked: 0 emails remaining.", "ERROR")
                dlg = ActivationDialog(self)
                dlg.exec()
                if not LicenseManager.can_send_email(1):
                    return
            else:
                self._signals.log.emit(f"Broadcast of {count} emails clamped to remaining limit of {status['remaining']}.", "WARNING")
                recs = recs[:status['remaining']]
                from qfluentwidgets import InfoBar, InfoBarPosition
                InfoBar.warning(
                    title="Trial Limit Active",
                    content=f"Your batch has been automatically clamped to your remaining limit of {status['remaining']} emails. Upgrade to Pro for unlimited broadcasts.",
                    orient=Qt.Horizontal,
                    isClosable=True,
                    position=InfoBarPosition.TOP,
                    duration=6000,
                    parent=self.window()
                )
                
        self._log("Initiating global broadcast...")
        self._set_sending_state(True)
        threading.Thread(target=self._worker_send, args=(recs,), daemon=True).start()

    def _set_sending_state(self, sending: bool):

        self._sending = sending
        self._stop_requested = False
        self._btn_send_all.setEnabled(not sending)
        self._btn_send_one.setEnabled(not sending)
        self._btn_stop.setEnabled(sending)

        self._status_badge.setText("SENDING" if sending else "READY")

    def _log(self, text: str, level: str = "INFO"):
        self._signals.log.emit(text, level)

    def _worker_send(self, recipients: list[str]):
        self._error_vault.clear()
        sent = 0; failed = 0; total = len(recipients)
        fresh_per_message = self._should_use_fresh_connection_per_message()
        try: 
            raw_val = int(self._interval_input.text().strip())
            interval = max(15, raw_val)
            if raw_val < 15:
                self._log(f"Enforcing minimum interval of 15 seconds (was {raw_val}s).", "WARNING")
        except: 
            interval = 30

        def _sleep_interruptibly(duration: float):
            deadline = time.monotonic() + max(duration, 0.0)
            while not self._stop_requested:
                remaining = deadline - time.monotonic()
                if remaining <= 0:
                    break
                time.sleep(min(remaining, 0.2))

        def _sleep_remaining(cycle_started_at: float):
            if self._stop_requested:
                return
            remaining = max(interval - (time.monotonic() - cycle_started_at), 0.0)
            if remaining > 0:
                _sleep_interruptibly(remaining)

        if fresh_per_message:
            self._log("Gmail mode detected. A fresh SMTP session will be opened for each recipient.", "INFO")
        else:
            try:
                self._log("Initializing SMTP Handshake...")
                last_err = None
                for attempt in range(1, 4):  # 3 retries with exponential backoff
                    try:
                        self._active_server = self._create_smtp_connection()
                        last_err = None
                        break
                    except Exception as e:
                        last_err = e
                        if self._stop_requested:
                            break
                        if attempt < 3:
                            wait = 2 ** attempt  # 2s, 4s
                            self._log(f"Handshake attempt {attempt}/3 failed: {e}. Retrying in {wait}s...", "WARNING")
                            _sleep_interruptibly(wait)
                
                if last_err:
                    raise last_err
                    
                if self._stop_requested:
                    try: self._active_server.quit()
                    except: pass
                    self._active_server = None
                    self._signals.finished.emit(0, 0, True)
                    return
                self._log("Authenticated successfully.", "SUCCESS")
            except Exception as e:
                if self._stop_requested:
                    self._signals.finished.emit(0, 0, True)
                else:
                    self._signals.error.emit(f"Handshake Failed after 3 attempts: {e}")
                    self._signals.finished.emit(0, 0, False)
                return

        # Filter already sent recipients before applying batch size
        message_signature = self._message_signature()
        filtered_recipients = []
        skipped_count = 0
        for rec in recipients:
            if self._was_message_sent(rec, message_signature):
                skipped_count += 1
            else:
                filtered_recipients.append(rec)
                
        recipients = filtered_recipients
        if skipped_count > 0:
            self._log(f"Skipped {skipped_count} recipient(s) (already sent).", "WARNING")
            
        if not recipients:
            self._log("No new recipients to email. Broadcast finished.", "SUCCESS")
            self._signals.finished.emit(0, 0, True)
            return

        # Respect Batch Size
        try: 
            bs = int(self._batch_size.text().strip())
            if bs < len(recipients):
                self._log(f"Limiting broadcast to batch size of {bs}...", "WARNING")
                recipients = recipients[:bs]
        except: pass

        total = len(recipients)

        for i, rec in enumerate(recipients):
            if self._stop_requested: break
            cycle_started_at = time.monotonic()
            
            # ── Robust Sending Logic ──
            try:
                msg = self._build_message(rec)
                if self._stop_requested: break

                if fresh_per_message:
                    self._reconnect_smtp(delay_seconds=0.35 if i > 0 else 0.0, silent=True)
                
                # Proactively ensure connection BEFORE sending
                try:
                    self._ensure_smtp_connection()
                except (smtplib.SMTPServerDisconnected, socket.error, ssl.SSLError):
                    self._log(f"Connection lost. Reconnecting for {rec}...", "WARNING")
                    try:
                        self._reconnect_smtp(delay_seconds=2.0)
                        self._log("Reconnected successfully.", "SUCCESS")
                    except Exception as reconn_err:
                        self._log(f"Reconnection failed for {rec}: {reconn_err}", "ERROR")
                        failed += 1
                        self._error_vault[rec] = f"Reconnection failed: {reconn_err}"
                        self._signals.progress.emit(i+1, total)
                        if i < total - 1 and not self._stop_requested:
                            _sleep_remaining(cycle_started_at)
                        continue

                # Actual Transmission
                try:
                    self._send_message_with_recovery(msg)
                except (smtplib.SMTPServerDisconnected, socket.error, ssl.SSLError):
                    self._log(f"Connection dropped during transmission. Retrying {rec}...", "WARNING")
                    try:
                        self._reconnect_smtp(delay_seconds=2.0)
                        self._log("Reconnected successfully. Retrying transmission...", "SUCCESS")
                        self._send_message_with_recovery(msg)
                    except Exception as retry_err:
                        self._log(f"Retry failed for {rec}: {retry_err}", "ERROR")
                        failed += 1
                        self._error_vault[rec] = f"Retry failed: {retry_err}"
                        self._active_server = None  # Force re-init on next loop
                        self._signals.progress.emit(i+1, total)
                        if i < total - 1 and not self._stop_requested:
                            _sleep_remaining(cycle_started_at)
                        continue
                except smtplib.SMTPResponseException as rate_err:
                    if rate_err.smtp_code in (421, 452):
                        self._log(f"Rate-limited by server (code {rate_err.smtp_code}). Waiting 60s before retrying {rec}...", "WARNING")
                        _sleep_interruptibly(60)
                        if self._stop_requested:
                            break
                        try:
                            self._reconnect_smtp()
                            self._send_message_with_recovery(msg)
                        except Exception as rl_err:
                            failed += 1
                            self._error_vault[rec] = f"Rate-limit retry failed: {rl_err}"
                            self._active_server = None
                            self._signals.progress.emit(i+1, total)
                            if i < total - 1 and not self._stop_requested:
                                _sleep_remaining(cycle_started_at)
                            continue
                    else:
                        raise  # Re-raise non-rate-limit SMTP errors
                    
                sent += 1
                from ..core.security import LicenseManager
                LicenseManager.record_email_send(1)
                self._record_successful_send(rec, message_signature)
                self._refresh_recipient_history_state()
                    
                self._log(f"✓ Sent to {rec} ({i+1}/{total})", "SUCCESS")

            except Exception as e:
                if self._stop_requested: break
                
                failed += 1
                self._error_vault[rec] = str(e)
                err_short = str(e).splitlines()[0][:60] + "..." if len(str(e)) > 60 else str(e)
                
                if "server not connected" in err_short.lower() or "closed" in err_short.lower():
                    self._log(f"Connection error for {rec}: {err_short}. Re-initializing for next recipient.", "ERROR")
                    self._active_server = None
                else:
                    self._on_log(f"Error for {rec}: <a href='err:{rec}' style='color: #FF453A; text-decoration: underline;'>{err_short} (Details)</a>", "ERROR")
            
            self._signals.progress.emit(i+1, total)
            if i < total - 1 and not self._stop_requested:
                _sleep_remaining(cycle_started_at)

        try: 
            if self._active_server:
                self._active_server.quit()
        except: pass
        self._active_server = None
        self._signals.finished.emit(sent, failed, self._stop_requested)



    def _salutation(self, contact_person: str | None) -> str:
        name  = (contact_person or "").strip()
        if not name:
            return "Sehr geehrte Damen und Herren,"
        first = name.split()[0].strip(" ,.;:").lower()
        import re
        if first in {"frau", "ms", "mrs"}:
            clean = re.sub(r"^(frau|ms|mrs)\s+", "", name, flags=re.IGNORECASE).strip()
            return f"Sehr geehrte Frau {clean},"
        if first in {"herr", "mr"}:
            clean = re.sub(r"^(herr|mr)\s+", "", name, flags=re.IGNORECASE).strip()
            return f"Sehr geehrter Herr {clean},"
        return f"Guten Tag {name},"

    def _build_message(self, recipient: str) -> MIMEMultipart:
        msg = MIMEMultipart()
        msg["From"] = f"{self._from_name.text()} <{self._smtp_user.text().strip()}>"
        
        reply_addr = self._reply_to.text().strip()
        if reply_addr:
            msg["Reply-To"] = reply_addr
            
        msg["To"] = recipient
        msg["Subject"] = self._subject.text()
        
        body_text = self._body_text.toPlainText()
        
        import sqlite3
        import re
        import os
        from ..core.config import get_memory_db_path, get_exports_dir
        
        anrede = "Sehr geehrte Damen und Herren,"
        company = "Firma"
        job_title = "Ausbildung"
        sender_name = ""
        sender_beruf = ""
        try:
            conn = sqlite3.connect(str(get_memory_db_path()), timeout=10.0)
            row = conn.execute("SELECT contact_person, company_name, job_title FROM leads WHERE email = ? LIMIT 1", (recipient,)).fetchone()
            if row:
                anrede = self._salutation(row[0])
                if row[1]: company = row[1]
                if row[2]: job_title = row[2]
                
            sender_row = conn.execute("SELECT value FROM settings WHERE key = 'sender_name'").fetchone()
            if sender_row and sender_row[0]:
                sender_name = sender_row[0]
                
            beruf_row = conn.execute("SELECT value FROM settings WHERE key = 'sender_beruf'").fetchone()
            if beruf_row and beruf_row[0]:
                sender_beruf = beruf_row[0]
            conn.close()
        except Exception:
            pass
            
        if not sender_name:
            from ..core.config import config_manager
            sender_name = config_manager.settings.email_from_name
            
        if not sender_name:
            sender_name = "Bewerber"
            
        body_text = body_text.replace("{{ANREDE}}", anrede)
        body_text = body_text.replace("{{COMPANY}}", company)
        
        msg.attach(MIMEText(body_text, "html" if self._type_toggle.isChecked() else "plain"))

        # Attach standard files
        for file_path in self._attachments:
            if not os.path.exists(file_path):
                continue
            
            try:
                part = MIMEBase("application", "octet-stream")
                with open(file_path, "rb") as f:
                    part.set_payload(f.read())
                
                encoders.encode_base64(part)
                filename = os.path.basename(file_path)
                part.add_header("Content-Disposition", "attachment", filename=filename)
                msg.attach(part)
            except Exception as e:
                self._log(f"Failed to attach {file_path}: {e}", "ERROR")

        # Attach dynamically generated lead PDF
        def sanitize(v):
            return re.sub(r'[<>:"/\\|?*]', '_', v)
            
        beruf_san = sanitize(sender_beruf or job_title or "Ausbildung")
        firma_san = sanitize(company)
        sender_san = sanitize(sender_name)
        
        pdf_filename = f"Bewerbung als {beruf_san} - {sender_san} @ {firma_san}.pdf"
        generic_pdf_filename = f"Bewerbung als {beruf_san} - {sender_san}.pdf"
        
        dynamic_pdf_path = get_exports_dir() / pdf_filename
        generic_pdf_path = get_exports_dir() / generic_pdf_filename
        raw_pdf_path = get_exports_dir() / "Bewerbung_Raw_Uploaded.pdf"
        
        if dynamic_pdf_path.exists():
            try:
                part = MIMEBase("application", "pdf")
                with open(dynamic_pdf_path, "rb") as f:
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", "attachment", filename=pdf_filename)
                msg.attach(part)
            except Exception as e:
                raise RuntimeError(f"Failed to attach generated PDF {pdf_filename}: {e}")
        elif generic_pdf_path.exists():
            try:
                part = MIMEBase("application", "pdf")
                with open(generic_pdf_path, "rb") as f:
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", "attachment", filename=generic_pdf_filename)
                msg.attach(part)
            except Exception as e:
                raise RuntimeError(f"Failed to attach generic PDF {generic_pdf_filename}: {e}")
        elif raw_pdf_path.exists():
            try:
                part = MIMEBase("application", "pdf")
                with open(raw_pdf_path, "rb") as f:
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", "attachment", filename=generic_pdf_filename) # Use generic filename to look professional
                msg.attach(part)
            except Exception as e:
                raise RuntimeError(f"Failed to attach raw PDF {raw_pdf_path}: {e}")
        else:
            if not self._attachments:
                raise FileNotFoundError(f"No PDF found. Tried dynamic ({pdf_filename}), generic ({generic_pdf_filename}), and raw ({raw_pdf_path}). Please generate it in the Edit Page first, or manually attach a CV.")

        return msg

    def _create_smtp_connection(self, silent: bool = False):
        from ..core.config import config_manager
        s = config_manager.settings

        # Prefer config values, fall back to any UI fields still present
        host = (getattr(s, "email_smtp_host", "") or self._smtp_host.text()).strip()
        port_txt = (getattr(s, "email_smtp_port", "587") or self._smtp_port.text()).strip()
        user = (getattr(s, "email_smtp_user", "") or self._smtp_user.text()).strip()
        pwd  = (getattr(s, "email_smtp_pass", "") or self._smtp_pass.text()).strip()

        if not host:
            raise ValueError(
                "SMTP Host not configured. Go to Settings → Email to enter your SMTP server."
            )
        try:
            port = int(port_txt)
        except ValueError:
            raise ValueError(f"Invalid SMTP Port: '{port_txt}'. Must be a number.")

        timeout = 300
        # Auto-select protocol: port 465 = Implicit SSL; everything else = plain+STARTTLS
        use_implicit_ssl = (port == 465)
        try:
            if use_implicit_ssl:
                if not silent:
                    self._signals.log.emit(f"Connecting via Implicit SSL to {host}:{port}...", "INFO")
                server = smtplib.SMTP_SSL(host, port, context=self._build_ssl_context(), timeout=timeout)
                server.ehlo()
            else:
                if not silent:
                    self._signals.log.emit(f"Connecting to {host}:{port} with STARTTLS...", "INFO")
                server = smtplib.SMTP(host, port, timeout=timeout)
                server.ehlo()
                server.starttls(context=self._build_ssl_context())
                server.ehlo()

            if user:
                if not silent:
                    self._signals.log.emit(f"Authenticating as {user}...", "INFO")
                server.login(user, pwd)

            return server

        except socket.timeout:
            raise Exception("Connection timed out. Check your host/port and firewall settings.")
        except ssl.SSLError as e:
            raise Exception(f"SSL/TLS Handshake failed: {e}. For Gmail use port 587 (STARTTLS) or 465 (SSL).")
        except ConnectionRefusedError:
            raise Exception("Connection refused. Ensure the SMTP server is reachable and the port is open.")
        except socket.gaierror as e:
            raise Exception(f"DNS resolution failed for '{host}': {e}. Check your internet connection and server hostname.")
        except OSError as e:
            raise Exception(f"Network error: {e}. Check your internet connection.")
        except smtplib.SMTPConnectError as e:
            raise Exception(f"Failed to connect to SMTP server: {e}")
        except smtplib.SMTPAuthenticationError as e:
            raise Exception(f"Authentication failed: {e}")
        except Exception as e:
            err_msg = str(e) or type(e).__name__
            raise Exception(err_msg)

    def _close_active_server(self):
        server = getattr(self, "_active_server", None)
        self._active_server = None
        if not server:
            return
        try:
            server.quit()
            return
        except Exception:
            pass
        try:
            server.close()
        except Exception:
            pass

    def _reconnect_smtp(self, delay_seconds: float = 0.8, silent: bool = False):
        self._close_active_server()
        if delay_seconds > 0:
            deadline = time.monotonic() + delay_seconds
            while not self._stop_requested:
                remaining = deadline - time.monotonic()
                if remaining <= 0:
                    break
                time.sleep(min(remaining, 0.2))
            if self._stop_requested:
                raise smtplib.SMTPServerDisconnected("Reconnect cancelled by user")
        self._active_server = self._create_smtp_connection(silent=silent)
        return self._active_server

    def _should_use_fresh_connection_per_message(self) -> bool:
        from ..core.config import config_manager
        s = config_manager.settings
        host = (getattr(s, "email_smtp_host", "") or self._smtp_host.text()).strip().lower()
        return host in {"smtp.gmail.com", "smtp.googlemail.com"}

    def _send_message_with_recovery(self, msg):
        if not self._active_server:
            raise smtplib.SMTPServerDisconnected("Server not connected")
        try:
            self._active_server.send_message(msg)
            return
        except smtplib.SMTPServerDisconnected:
            raise
        except (socket.error, ssl.SSLError):
            raise
        except OSError as e:
            # Some Windows 10 failures surface as generic socket/transport errors.
            lower = str(e).lower()
            if any(token in lower for token in ("not connected", "connection reset", "forcibly closed", "broken pipe", "timed out")):
                raise smtplib.SMTPServerDisconnected(str(e))
            raise

    def _ensure_smtp_connection(self):
        """Verify the connection is still alive, otherwise raise SMTPServerDisconnected."""
        if not self._active_server:
            raise smtplib.SMTPServerDisconnected("No active connection")
        try:
            # Use a shorter timeout for the liveness check
            old_timeout = self._active_server.timeout
            self._active_server.timeout = 10
            status = self._active_server.noop()
            self._active_server.timeout = old_timeout
            if not (isinstance(status, tuple) and status[0] and 200 <= int(status[0]) < 400):
                raise smtplib.SMTPServerDisconnected(f"SMTP session invalid (status {status[0]})")
        except (smtplib.SMTPServerDisconnected, socket.error, ssl.SSLError) as e:
            raise smtplib.SMTPServerDisconnected(str(e))
        except Exception as e:
            raise smtplib.SMTPServerDisconnected(f"Session check failed: {e}")

    def _test_connection(self):
        """Run a one-off connection test in a background thread."""
        self._on_log("Testing SMTP connection...", "INFO")
        
        def run_test():
            try:
                server = self._create_smtp_connection()
                try: server.quit()
                except: pass
                # Using lambda to avoid direct UI calls from thread
                self._signals.log.emit("Connection test successful! Your credentials are valid.", "SUCCESS")
                from .event_bridge import event_bus
                event_bus.emit("toast.show", title="SMTP Success", subtitle="Connection verified successfully.", type="success")
            except Exception as e:
                self._signals.error.emit(f"Connection test failed: {e}")
                from .event_bridge import event_bus
                event_bus.emit("toast.show", title="SMTP Failed", subtitle=str(e), type="error")

        threading.Thread(target=run_test, daemon=True).start()

    def _save_fields(self):
        if getattr(self, '_is_restoring', False):
            return
        self._remember_current_sender_profile()
        current_settings = config_manager.settings
        smtp_host = self._smtp_host.text().strip() or getattr(current_settings, "email_smtp_host", "") or "smtp.gmail.com"
        smtp_port = self._smtp_port.text().strip() or getattr(current_settings, "email_smtp_port", "") or "587"
        config_manager.update(
            email_smtp_host=smtp_host,
            email_smtp_port=smtp_port,
            email_smtp_user=self._smtp_user.text(),
            email_smtp_pass=self._smtp_pass.text(),
            email_sender_profiles=list(self._sender_profiles.values()),
            email_from_name=self._from_name.text(),
            email_reply_to=self._reply_to.text(),
            email_smtp_auth=self._auth_switch.isChecked(),
            email_smtp_ssl=self._ssl_switch.isChecked(),
            email_smtp_tls=self._tls_switch.isChecked(),
            email_body_html=self._type_toggle.isChecked(),
            email_subject=self._subject.text(),
            email_body=self._body_text.toPlainText(),
            email_interval=self._interval_input.text(),
            email_recipients="\n".join(self._get_recipients()),
            email_attachments="\n".join(self._attachments),
        )

    def _restore_fields(self):
        s = config_manager.settings
        self._load_sender_profiles()
        self._smtp_host.setText(s.email_smtp_host or "smtp.gmail.com")
        self._smtp_port.setText(s.email_smtp_port or "587")
        self._smtp_user.setText(s.email_smtp_user)
        self._smtp_pass.setText(s.email_smtp_pass)
        self._from_name.setText(s.email_from_name)
        self._reply_to.setText(s.email_reply_to)
        self._auth_switch.setChecked(s.email_smtp_auth)
        self._ssl_switch.setChecked(s.email_smtp_ssl)
        self._tls_switch.setChecked(s.email_smtp_tls)
        self._type_toggle.setChecked(s.email_body_html)
        self._subject.setText(s.email_subject)
        
        if s.email_body:
            self._body_text.setPlainText(s.email_body)
        else:
            self._body_text.setPlainText(
                "{{ANREDE}}\n\n"
                "gerne möchte ich mich um den von Ihnen ausgeschriebenen Ausbildungsplatz bewerben. "
                "Im Anhang finden Sie meine Bewerbungsunterlagen als PDF-Datei.\n\n"
                "Bei Rückfragen stehe ich Ihnen gerne zur Verfügung. Ich freue mich, von Ihnen zu hören!\n\n"
                "Mit freundlichen Grüßen"
            )
        self._interval_input.setText(s.email_interval)
        self._attachments = [
            path for path in (s.email_attachments or "").split("\n")
            if path.strip() and os.path.exists(path.strip())
        ]
        self._update_attachment_badge()
        if s.email_recipients:
            self._set_recipients(s.email_recipients.split("\n"))
        else:
            self._on_recipients_changed()

    def _on_recipients_changed(self):
        """Updates the numeric badge and persists the updated list."""
        count = self._recipient_list.count()
        self._rec_count.setText(f"{count} EMAIL(S)")
        if count == 0:
            self._rec_stack.setCurrentWidget(self._rec_empty)
        else:
            self._rec_stack.setCurrentWidget(self._recipient_list)
        self._save_fields()


    # End of class
